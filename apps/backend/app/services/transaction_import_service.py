from __future__ import annotations

import csv
import io
import json
import logging
import re
import unicodedata
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.llm.runtime import build_llm_runtime
from app.llm.types import LlmObservabilityClient
from app.models.account import Account
from app.models.category import Category
from app.repositories.account_repository import AccountRepository
from app.repositories.category_repository import CategoryRepository
from app.repositories.transaction_repository import TransactionRepository
from app.schemas.transactions import (
    TransactionCategoryAssistantDraft,
    TransactionCategoryAssistantSuggestion,
    TransactionCreate,
    TransactionImportAnalysisResponse,
    TransactionImportColumnMapping,
    TransactionImportCommitItem,
    TransactionImportCommitRequest,
    TransactionImportCommitResponse,
    TransactionImportDraft,
    TransactionImportPreviewResponse,
)
from app.services.azure_document_intelligence_ocr_service import (
    AzureDocumentIntelligenceOcrService,
)
from app.services.azure_openai_pdf_parser_service import AzureOpenAIPdfParserService
from app.services.azure_openai_transaction_category_service import (
    AzureOpenAITransactionCategoryService,
)

SUPPORTED_IMPORT_TYPES = {"csv", "xlsx", "xlsm", "xltx", "xltm", "pdf"}
REQUIRED_IMPORT_FIELDS = ("date", "amount", "description")
PDF_EXTRACTED_COLUMNS = ["Fecha", "Descripción", "Importe"]

FIELD_ALIASES = {
    "date": {
        "date",
        "fecha",
        "fechainicio",
        "fechadeinicio",
        "startdate",
        "transactiondate",
        "bookingdate",
        "posteddate",
        "valuedate",
    },
    "amount": {"amount", "importe", "total", "value", "sum", "quantity"},
    "description": {
        "description",
        "descripcion",
        "concept",
        "concepto",
        "merchant",
        "payee",
        "beneficiary",
        "details",
    },
    "category": {"category", "categoria"},
    "notes": {"notes", "note", "nota", "comment", "comments", "memo", "reference"},
}

DESCRIPTION_STOPWORDS = {
    "card",
    "compra",
    "debito",
    "debit",
    "pago",
    "payment",
    "pos",
    "purchase",
    "ref",
    "reference",
    "sepa",
    "tarjeta",
    "transfer",
    "transferencia",
    "trx",
    "visa",
}

ASSISTED_CLASSIFICATION_CONFIDENCE_THRESHOLD = 0.5

logger = logging.getLogger(__name__)


@dataclass
class ParsedImportFile:
    source_type: str
    columns: list[str]
    rows: list[dict[str, str]]
    message: str | None = None


@dataclass
class CategorySuggestion:
    source_row_number: int
    category_id: uuid.UUID
    label: str
    source: str
    confidence: float | None
    reason: str
    model: str | None = None


@dataclass
class CategoryHistoryIndex:
    exact_matches: dict[str, Counter[uuid.UUID]]
    merchant_patterns: dict[str, Counter[uuid.UUID]]


@dataclass(frozen=True)
class TransactionImportFingerprint:
    account_id: uuid.UUID
    date: date
    amount: Decimal
    currency: str
    description: str
    notes: str


class TransactionImportService:
    def __init__(
        self,
        repository: TransactionRepository,
        account_repository: AccountRepository,
        category_repository: CategoryRepository,
        db: Session,
        document_ocr_service: AzureDocumentIntelligenceOcrService | None = None,
        pdf_llm_parser_service: AzureOpenAIPdfParserService | None = None,
        category_classifier_service: AzureOpenAITransactionCategoryService | None = None,
        llm_observability_client: LlmObservabilityClient | None = None,
    ) -> None:
        self.repository = repository
        self.account_repository = account_repository
        self.category_repository = category_repository
        self.db = db
        llm_runtime = build_llm_runtime()
        shared_prompt_provider = llm_runtime.prompt_provider
        shared_observability_client = llm_observability_client or llm_runtime.observability_client
        self.document_ocr_service = document_ocr_service or AzureDocumentIntelligenceOcrService()
        self.pdf_llm_parser_service = pdf_llm_parser_service or AzureOpenAIPdfParserService(
            prompt_provider=shared_prompt_provider,
            observability_client=shared_observability_client,
        )
        self.category_classifier_service = (
            category_classifier_service
            or AzureOpenAITransactionCategoryService(
                prompt_provider=shared_prompt_provider,
                observability_client=shared_observability_client,
            )
        )
        self.llm_observability_client = shared_observability_client

    async def analyze_file(self, *, file: UploadFile) -> TransactionImportAnalysisResponse:
        parsed = await self._parse_upload(file, user_id=None)
        return TransactionImportAnalysisResponse(
            source_type=parsed.source_type,
            columns=parsed.columns,
            sample_rows=parsed.rows[:5],
            suggested_mapping=self._suggest_mapping(parsed.columns),
            total_rows=len(parsed.rows),
            message=parsed.message,
        )

    async def build_preview(
        self,
        *,
        user_id: uuid.UUID,
        account_id: uuid.UUID,
        file: UploadFile,
        mapping_json: str,
    ) -> TransactionImportPreviewResponse:
        account = self._require_account(user_id=user_id, account_id=account_id)
        parsed = await self._parse_upload(file, user_id=user_id)

        mapping = self._build_effective_mapping(parsed.source_type, mapping_json)
        self._validate_mapping(mapping)

        categories = self.category_repository.list_all_for_user(
            user_id=user_id,
            sort_by="name",
            sort_order="asc",
        )
        category_by_name = {
            self._normalize_column_name(category.name): category for category in categories
        }

        rows = [
            self._build_draft(
                raw_row=raw_row,
                row_number=index + 1,
                account=account,
                mapping=mapping,
                category_by_name=category_by_name,
            )
            for index, raw_row in enumerate(parsed.rows)
        ]
        rows = self._classify_drafts(
            user_id=user_id,
            rows=rows,
            categories=categories,
        )

        return TransactionImportPreviewResponse(
            source_type=parsed.source_type,
            account_id=account.id,
            account_currency=account.currency,
            imported_count=len(rows),
            rows=rows,
        )

    def commit_import(
        self,
        *,
        user_id: uuid.UUID,
        payload: TransactionImportCommitRequest,
    ) -> TransactionImportCommitResponse:
        existing_fingerprints = self._existing_import_fingerprints(
            user_id=user_id,
            items=payload.items,
        )
        seen_fingerprints = set(existing_fingerprints)
        imported_count = 0
        skipped_duplicates = 0

        for item in payload.items:
            account = self._require_account(user_id=user_id, account_id=item.account_id)
            category = self._require_category_if_present(
                user_id=user_id,
                category_id=item.category_id,
            )
            if item.currency != account.currency:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Transaction currency must match the selected account currency",
                )
            if category is not None:
                # Category compatibility is intentionally permissive in this v1.
                pass

            fingerprint = self._transaction_fingerprint(
                account_id=item.account_id,
                date_value=item.date,
                amount=item.amount,
                currency=item.currency,
                description=item.description,
                notes=item.notes,
            )
            if fingerprint in seen_fingerprints:
                skipped_duplicates += 1
                continue

            create_payload = TransactionCreate(
                account_id=item.account_id,
                category_id=item.category_id,
                date=item.date,
                amount=item.amount,
                currency=item.currency,
                description=item.description,
                notes=item.notes,
            )
            self.repository.create(user_id=user_id, payload=create_payload.model_dump())
            seen_fingerprints.add(fingerprint)
            imported_count += 1

        self.db.commit()
        return TransactionImportCommitResponse(
            imported_count=imported_count,
            skipped_duplicates=skipped_duplicates,
        )

    async def _parse_upload(
        self,
        file: UploadFile,
        *,
        user_id: uuid.UUID | None,
    ) -> ParsedImportFile:
        extension = Path(file.filename or "").suffix.lower().lstrip(".")
        if extension not in SUPPORTED_IMPORT_TYPES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Supported import formats are CSV, Excel and PDF",
            )

        content = await file.read()
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded file is empty",
            )

        if extension == "csv":
            return self._parse_csv(content)
        if extension in {"xlsx", "xlsm", "xltx", "xltm"}:
            return self._parse_excel(content)
        return self._parse_pdf(content, user_id=user_id)

    def _parse_csv(self, content: bytes) -> ParsedImportFile:
        text = self._decode_csv_content(content)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        columns = [
            column.strip() for column in (reader.fieldnames or []) if column and column.strip()
        ]
        rows = []
        for row in reader:
            normalized = {
                key.strip(): (value.strip() if isinstance(value, str) else "")
                for key, value in row.items()
                if key and key.strip()
            }
            if any(value for value in normalized.values()):
                rows.append(normalized)

        return ParsedImportFile(source_type="csv", columns=columns, rows=rows)

    def _parse_excel(self, content: bytes) -> ParsedImportFile:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except (BadZipFile, OSError, ValueError) as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded spreadsheet is not a valid Excel workbook",
            ) from exc
        sheet = workbook.active
        if sheet is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded spreadsheet does not contain a readable sheet",
            )
        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded spreadsheet is empty",
            ) from exc

        columns = [
            str(value).strip() for value in header_row if value is not None and str(value).strip()
        ]
        parsed_rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            row_map: dict[str, str] = {}
            for index, column in enumerate(columns):
                value = raw_row[index] if index < len(raw_row) else None
                row_map[column] = self._to_cell_string(value)
            if any(value for value in row_map.values()):
                parsed_rows.append(row_map)

        return ParsedImportFile(source_type="excel", columns=columns, rows=parsed_rows)

    def _parse_pdf(self, content: bytes, *, user_id: uuid.UUID | None) -> ParsedImportFile:
        flow = self.llm_observability_client.start_flow(
            "pdf_transaction_parser_flow",
            input_payload={"content_bytes": len(content)},
            metadata={
                "user_id": str(user_id) if user_id is not None else None,
                "source_type": "pdf",
            },
        )
        try:
            ocr_result = self.document_ocr_service.extract_text(content=content)
            rows = self._extract_rows_from_pdf_structured_text(ocr_result.structured_text)
            if rows:
                self.llm_observability_client.end_flow(
                    flow,
                    output_payload={"row_count": len(rows)},
                    metadata={
                        "fallback_used": False,
                        "structured_table_match": True,
                    },
                )
                return ParsedImportFile(
                    source_type="pdf",
                    columns=PDF_EXTRACTED_COLUMNS,
                    rows=rows,
                    message=(
                        "Hemos leído el PDF con OCR, generado un texto estructurado y preparado "
                        "una propuesta de transacciones. Revísala antes de confirmar "
                        "la importación."
                    ),
                )

            llm_rows = self.pdf_llm_parser_service.parse_transactions(
                structured_text=ocr_result.structured_text,
                tables_markdown=ocr_result.tables_markdown,
            )
            if llm_rows:
                self.llm_observability_client.end_flow(
                    flow,
                    output_payload={"row_count": len(llm_rows)},
                    metadata={
                        "fallback_used": True,
                        "structured_table_match": False,
                    },
                )
                return ParsedImportFile(
                    source_type="pdf",
                    columns=PDF_EXTRACTED_COLUMNS,
                    rows=llm_rows,
                    message=(
                        "Hemos leído el PDF con OCR y usado una capa asistida para interpretar "
                        "las transacciones antes de la revisión manual."
                    ),
                )

            self.llm_observability_client.end_flow(
                flow,
                output_payload={"row_count": 0},
                metadata={
                    "fallback_used": True,
                    "structured_table_match": False,
                },
                status_message="No transaction rows could be extracted",
                level="WARNING",
            )
            return ParsedImportFile(
                source_type="pdf",
                columns=PDF_EXTRACTED_COLUMNS,
                rows=[],
                message=(
                    "Hemos leído el PDF con OCR, pero no hemos podido convertirlo en movimientos "
                    "revisables todavía. Prueba con otro extracto o utiliza CSV/Excel."
                ),
            )
        except Exception as exc:
            self.llm_observability_client.end_flow(
                flow,
                output_payload=None,
                metadata={
                    "fallback_used": True,
                    "structured_table_match": False,
                },
                status_message=str(exc),
                level="ERROR",
            )
            raise

    def _suggest_mapping(self, columns: list[str]) -> TransactionImportColumnMapping:
        suggestions: dict[str, str | None] = {field: None for field in FIELD_ALIASES}

        for column in columns:
            normalized_column = self._normalize_column_name(column)
            for field, aliases in FIELD_ALIASES.items():
                if suggestions[field] is None and normalized_column in aliases:
                    suggestions[field] = column

        return TransactionImportColumnMapping(**suggestions)

    def _parse_mapping(self, raw_mapping: str) -> TransactionImportColumnMapping:
        try:
            payload = json.loads(raw_mapping)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Import mapping is not valid JSON",
            ) from exc

        try:
            return TransactionImportColumnMapping.model_validate(payload)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Import mapping is invalid",
            ) from exc

    def _build_effective_mapping(
        self,
        source_type: str,
        raw_mapping: str,
    ) -> TransactionImportColumnMapping:
        if source_type != "pdf":
            return self._parse_mapping(raw_mapping)

        return TransactionImportColumnMapping(
            date="Fecha",
            amount="Importe",
            description="Descripción",
            category=None,
            notes=None,
        )

    def _validate_mapping(self, mapping: TransactionImportColumnMapping) -> None:
        missing_fields = [field for field in REQUIRED_IMPORT_FIELDS if not getattr(mapping, field)]
        if missing_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required import fields: {', '.join(missing_fields)}",
            )

    def _build_draft(
        self,
        *,
        raw_row: dict[str, str],
        row_number: int,
        account: Account,
        mapping: TransactionImportColumnMapping,
        category_by_name: dict[str, Category],
    ) -> TransactionImportDraft:
        description_raw = self._mapped_value(raw_row, mapping.description)
        notes_raw = self._mapped_value(raw_row, mapping.notes)
        category_raw = self._mapped_value(raw_row, mapping.category)
        raw_date = self._mapped_value(raw_row, mapping.date)

        parsed_date = (
            None
            if self._date_requires_manual_review(raw_date=raw_date, column_name=mapping.date)
            else self._parse_date(raw_date)
        )
        parsed_amount = self._parse_amount(self._mapped_value(raw_row, mapping.amount))

        validation_errors: list[str] = []
        if parsed_date is None:
            validation_errors.append("Review the date")
        if parsed_amount is None:
            validation_errors.append("Review the amount")
        if not description_raw:
            validation_errors.append("Review the description")

        matched_category = (
            category_by_name.get(self._normalize_column_name(category_raw))
            if category_raw
            else None
        )

        return TransactionImportDraft(
            source_row_number=row_number,
            account_id=account.id,
            category_id=matched_category.id if matched_category else None,
            category_label=category_raw or None,
            date=parsed_date,
            amount=parsed_amount,
            currency=account.currency,
            description=description_raw or None,
            notes=notes_raw or None,
            validation_errors=validation_errors,
        )

    def _classify_drafts(
        self,
        *,
        user_id: uuid.UUID,
        rows: list[TransactionImportDraft],
        categories: list[Category],
    ) -> list[TransactionImportDraft]:
        if not rows:
            return rows

        category_map = {category.id: category for category in categories}
        history = self._build_category_history_index(
            user_id=user_id,
            category_map=category_map,
        )

        classified_rows: list[TransactionImportDraft] = []
        assisted_candidates: list[TransactionImportDraft] = []

        for row in rows:
            if row.category_id is not None:
                final_row = self._attach_classification_debug(
                    row=row,
                    reason="Category came explicitly from the imported file",
                    model=None,
                )
                classified_rows.append(final_row)
                continue

            suggestion = self._suggest_category_from_history(
                row=row,
                history=history,
                category_map=category_map,
            )
            if suggestion is not None:
                final_row = self._apply_category_suggestion(row=row, suggestion=suggestion)
                classified_rows.append(final_row)
                continue

            final_row = self._attach_classification_debug(
                row=row,
                reason=(
                    "No confident match from history or repeated patterns"
                    if self.category_classifier_service.enabled
                    else (
                        "No confident match from history or repeated patterns, "
                        "and assisted layer is disabled"
                    )
                ),
                model=self.category_classifier_service.model_name
                if self.category_classifier_service.enabled
                else None,
            )
            classified_rows.append(final_row)
            if self._can_use_assisted_classification(row=row):
                assisted_candidates.append(final_row)

        if not assisted_candidates:
            for row in classified_rows:
                self._log_classification_event(row=row)
            return classified_rows

        assisted_suggestions = self._suggest_categories_with_assisted_layer(
            user_id=user_id,
            rows=assisted_candidates,
            categories=categories,
            category_map=category_map,
        )
        if not assisted_suggestions:
            for row in classified_rows:
                if row.category_id is None:
                    self._log_classification_event(row=row)
            return classified_rows

        suggestion_by_row = {
            suggestion.source_row_number: suggestion for suggestion in assisted_suggestions
        }
        final_rows: list[TransactionImportDraft] = []
        for row in classified_rows:
            if row.category_id is None and row.source_row_number in suggestion_by_row:
                final_row = self._apply_category_suggestion(
                    row=row,
                    suggestion=suggestion_by_row[row.source_row_number],
                )
            else:
                final_row = row
            self._log_classification_event(row=final_row)
            final_rows.append(final_row)
        return final_rows

    def _build_category_history_index(
        self,
        *,
        user_id: uuid.UUID,
        category_map: dict[uuid.UUID, Category],
    ) -> CategoryHistoryIndex:
        exact_matches: dict[str, Counter[uuid.UUID]] = defaultdict(Counter)
        merchant_patterns: dict[str, Counter[uuid.UUID]] = defaultdict(Counter)

        transactions = self.repository.list_all_for_user(
            user_id=user_id,
            sort_by="date",
            sort_order="desc",
        )
        for transaction in transactions:
            category_id = transaction.category_id
            if category_id is None or category_id not in category_map:
                continue

            description_key = self._normalize_description_key(transaction.description)
            if description_key:
                exact_matches[description_key][category_id] += 1

            merchant_key = self._merchant_pattern_key(transaction.description)
            if merchant_key:
                merchant_patterns[merchant_key][category_id] += 1

        return CategoryHistoryIndex(
            exact_matches=dict(exact_matches),
            merchant_patterns=dict(merchant_patterns),
        )

    def _suggest_category_from_history(
        self,
        *,
        row: TransactionImportDraft,
        history: CategoryHistoryIndex,
        category_map: dict[uuid.UUID, Category],
    ) -> CategorySuggestion | None:
        description = row.description or ""
        exact_key = self._normalize_description_key(description)
        merchant_key = self._merchant_pattern_key(description)
        allowed_category_ids = self._compatible_category_ids(
            amount=row.amount,
            category_map=category_map,
        )

        suggestion = self._pick_category_from_counter(
            source_row_number=row.source_row_number,
            counter=history.exact_matches.get(exact_key),
            category_map=category_map,
            allowed_category_ids=allowed_category_ids,
            min_ratio=0.75,
            source="history",
        )
        if suggestion is not None:
            return suggestion

        return self._pick_category_from_counter(
            source_row_number=row.source_row_number,
            counter=history.merchant_patterns.get(merchant_key),
            category_map=category_map,
            allowed_category_ids=allowed_category_ids,
            min_ratio=0.75,
            source="pattern",
        )

    def _pick_category_from_counter(
        self,
        *,
        source_row_number: int,
        counter: Counter[uuid.UUID] | None,
        category_map: dict[uuid.UUID, Category],
        allowed_category_ids: set[uuid.UUID],
        min_ratio: float,
        source: str,
    ) -> CategorySuggestion | None:
        if counter is None:
            return None

        compatible_counts = {
            category_id: count
            for category_id, count in counter.items()
            if category_id in allowed_category_ids
        }
        if not compatible_counts:
            return None

        ranked = sorted(
            compatible_counts.items(),
            key=lambda item: (-item[1], str(item[0])),
        )
        top_category_id, top_count = ranked[0]
        if len(ranked) > 1 and ranked[1][1] == top_count:
            return None

        total = sum(compatible_counts.values())
        ratio = top_count / total if total else 0
        if ratio < min_ratio:
            return None

        category = category_map.get(top_category_id)
        if category is None:
            return None

        confidence = round(min(0.99, 0.72 + (ratio * 0.18) + min(total, 4) * 0.03), 2)
        return CategorySuggestion(
            source_row_number=source_row_number,
            category_id=top_category_id,
            label=category.name,
            source=source,
            confidence=confidence,
            reason=(
                "Matched exact transaction history"
                if source == "history"
                else "Matched repeated merchant pattern"
            ),
        )

    def _compatible_category_ids(
        self,
        *,
        amount: Decimal | None,
        category_map: dict[uuid.UUID, Category],
    ) -> set[uuid.UUID]:
        if amount is None:
            return set(category_map.keys())

        if amount > 0:
            preferred_type = "income"
        elif amount < 0:
            preferred_type = "expense"
        else:
            preferred_type = None

        if preferred_type is None:
            return set(category_map.keys())

        return {
            category_id
            for category_id, category in category_map.items()
            if category.type.value in {preferred_type, "transfer"}
        }

    def _suggest_categories_with_assisted_layer(
        self,
        *,
        user_id: uuid.UUID,
        rows: list[TransactionImportDraft],
        categories: list[Category],
        category_map: dict[uuid.UUID, Category],
    ) -> list[CategorySuggestion]:
        if not rows or not self.category_classifier_service.enabled:
            return []

        assistant_rows = [
            TransactionCategoryAssistantDraft(
                source_row_number=row.source_row_number,
                description=row.description or "",
                notes=row.notes or None,
                amount=row.amount,
                currency=row.currency,
            )
            for row in rows
            if row.description and row.amount is not None
        ]
        if not assistant_rows:
            return []

        flow = self.llm_observability_client.start_flow(
            "transaction_category_assistant_flow",
            input_payload={
                "row_count": len(assistant_rows),
                "category_count": len(categories),
            },
            metadata={
                "user_id": str(user_id),
                "classification_threshold": ASSISTED_CLASSIFICATION_CONFIDENCE_THRESHOLD,
            },
        )
        try:
            raw_suggestions = self.category_classifier_service.classify_rows(
                rows=assistant_rows,
                categories=categories,
            )
            suggestions = [
                suggestion
                for raw_suggestion in raw_suggestions
                if (
                    suggestion := self._normalize_assisted_suggestion(
                        raw_suggestion=raw_suggestion,
                        row_by_number={row.source_row_number: row for row in rows},
                        category_map=category_map,
                    )
                )
                is not None
            ]
            self.llm_observability_client.end_flow(
                flow,
                output_payload={"suggestion_count": len(suggestions)},
                metadata={
                    "accepted_suggestion_count": len(suggestions),
                    "raw_suggestion_count": len(raw_suggestions),
                    "suggestion_source": "assisted",
                },
            )
            return suggestions
        except Exception as exc:
            self.llm_observability_client.end_flow(
                flow,
                output_payload=None,
                metadata={"suggestion_source": "assisted"},
                status_message=str(exc),
                level="ERROR",
            )
            raise

    def _normalize_assisted_suggestion(
        self,
        *,
        raw_suggestion: TransactionCategoryAssistantSuggestion,
        row_by_number: dict[int, TransactionImportDraft],
        category_map: dict[uuid.UUID, Category],
    ) -> CategorySuggestion | None:
        category_id = raw_suggestion.category_id
        confidence = raw_suggestion.confidence
        if (
            category_id is None
            or confidence is None
            or confidence < ASSISTED_CLASSIFICATION_CONFIDENCE_THRESHOLD
        ):
            return None

        category = category_map.get(category_id)
        row = row_by_number.get(raw_suggestion.source_row_number)
        if category is None or row is None:
            return None

        allowed_category_ids = self._compatible_category_ids(
            amount=row.amount,
            category_map=category_map,
        )
        if category_id not in allowed_category_ids:
            return None

        return CategorySuggestion(
            source_row_number=raw_suggestion.source_row_number,
            category_id=category_id,
            label=category.name,
            source="assisted",
            confidence=round(confidence, 2),
            reason=(f"Assistant suggested category with confidence {round(confidence, 2)}"),
            model=self.category_classifier_service.model_name,
        )

    def _apply_category_suggestion(
        self,
        *,
        row: TransactionImportDraft,
        suggestion: CategorySuggestion,
    ) -> TransactionImportDraft:
        return row.model_copy(
            update={
                "category_id": suggestion.category_id,
                "category_suggestion_label": suggestion.label,
                "category_suggestion_source": suggestion.source,
                "category_suggestion_confidence": suggestion.confidence,
                "category_suggestion_reason": (
                    suggestion.reason if settings.classification_debug else None
                ),
                "category_suggestion_model": (
                    suggestion.model if settings.classification_debug else None
                ),
                "category_is_suggested": True,
            }
        )

    def _attach_classification_debug(
        self,
        *,
        row: TransactionImportDraft,
        reason: str,
        model: str | None,
    ) -> TransactionImportDraft:
        if not settings.classification_debug:
            return row
        return row.model_copy(
            update={
                "category_suggestion_reason": reason,
                "category_suggestion_model": model,
            }
        )

    def _log_classification_event(self, *, row: TransactionImportDraft) -> None:
        if not settings.classification_debug:
            return

        payload = {
            "event": "transaction_import_classification",
            "source_row_number": row.source_row_number,
            "description": row.description,
            "amount": str(row.amount) if row.amount is not None else None,
            "category_id": str(row.category_id) if row.category_id is not None else None,
            "category_is_suggested": row.category_is_suggested,
            "category_suggestion_label": row.category_suggestion_label,
            "category_suggestion_source": row.category_suggestion_source,
            "category_suggestion_confidence": row.category_suggestion_confidence,
            "category_suggestion_reason": row.category_suggestion_reason,
            "category_suggestion_model": row.category_suggestion_model,
        }
        logger.info("%s", json.dumps(payload, ensure_ascii=False))

    def _can_use_assisted_classification(self, *, row: TransactionImportDraft) -> bool:
        return bool(row.description and row.amount is not None)

    def _extract_rows_from_pdf_structured_text(self, structured_text: str) -> list[dict[str, str]]:
        table_blocks = re.findall(
            r"\[Table \d+\]\n((?:\|.*\|\n?)*)",
            structured_text,
            flags=re.MULTILINE,
        )
        rows: list[dict[str, str]] = []
        for block in table_blocks:
            grid = self._markdown_table_to_grid(block)
            if not self._grid_looks_like_transaction_table(grid):
                continue
            rows.extend(self._transaction_rows_from_grid(grid))
        return rows

    def _markdown_table_to_grid(self, markdown_table: str) -> list[list[str]]:
        rows: list[list[str]] = []
        for raw_line in markdown_table.splitlines():
            line = raw_line.strip()
            if not line.startswith("|") or not line.endswith("|"):
                continue
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if cells and all(set(cell) <= {"-"} for cell in cells if cell):
                continue
            rows.append(cells)
        return rows

    def _grid_looks_like_transaction_table(self, grid: list[list[str]]) -> bool:
        if len(grid) < 2:
            return False
        headers = [self._normalize_column_name(value) for value in grid[0]]
        has_date = any("fecha" in header for header in headers)
        has_description = any("descripcion" in header for header in headers)
        has_money = any(
            "dinerosaliente" in header or "dineroentrante" in header or "importe" in header
            for header in headers
        )
        return has_date and has_description and has_money

    def _transaction_rows_from_grid(self, grid: list[list[str]]) -> list[dict[str, str]]:
        header_map = {
            self._normalize_column_name(value): index
            for index, value in enumerate(grid[0])
            if value.strip()
        }

        transaction_date_index = self._find_header_index(
            header_map,
            "fechadelatransaccion",
            "fecha",
        )
        description_index = self._find_header_index(header_map, "descripcion")
        outgoing_index = self._find_header_index(header_map, "dinerosaliente", "importe")
        incoming_index = self._find_header_index(header_map, "dineroentrante")

        if transaction_date_index is None or description_index is None:
            return []

        rows: list[dict[str, str]] = []
        for raw_row in grid[1:]:
            transaction_date = self._table_value(raw_row, transaction_date_index)
            description = self._table_value(raw_row, description_index)
            outgoing = self._table_value(raw_row, outgoing_index)
            incoming = self._table_value(raw_row, incoming_index)
            amount = self._signed_amount_from_columns(outgoing=outgoing, incoming=incoming)

            if not transaction_date or not description or not amount:
                continue

            rows.append(
                {
                    "Fecha": transaction_date,
                    "Descripción": description,
                    "Importe": amount,
                }
            )

        return rows

    def _find_header_index(self, header_map: dict[str, int], *candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in header_map:
                return header_map[candidate]
        return None

    def _table_value(self, row: list[str], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return row[index].strip()

    def _signed_amount_from_columns(self, *, outgoing: str, incoming: str) -> str:
        if outgoing:
            return self._ensure_amount_sign(outgoing, negative=True)
        if incoming:
            return self._ensure_amount_sign(incoming, negative=False)
        return ""

    def _ensure_amount_sign(self, value: str, *, negative: bool) -> str:
        cleaned = value.strip()
        if not cleaned:
            return ""
        if negative and not cleaned.startswith("-"):
            return f"-{cleaned.lstrip('+')}"
        if not negative and cleaned.startswith("-"):
            return cleaned.lstrip("-")
        return cleaned.lstrip("+")

    def _mapped_value(self, row: dict[str, str], column_name: str | None) -> str:
        if not column_name:
            return ""
        return (row.get(column_name) or "").strip()

    def _parse_date(self, value: str) -> date | None:
        if not value:
            return None

        cleaned = self._normalize_human_date(value.strip())
        for parser in (
            lambda current: date.fromisoformat(current),
            lambda current: datetime.fromisoformat(current.replace("Z", "+00:00")).date(),
            lambda current: datetime.strptime(current, "%d %m %Y").date(),
            lambda current: datetime.strptime(current, "%d/%m/%Y").date(),
            lambda current: datetime.strptime(current, "%d/%m/%Y %H:%M:%S").date(),
            lambda current: datetime.strptime(current, "%d/%m/%Y %H:%M").date(),
            lambda current: datetime.strptime(current, "%m/%d/%Y").date(),
            lambda current: datetime.strptime(current, "%m/%d/%Y %H:%M:%S").date(),
            lambda current: datetime.strptime(current, "%m/%d/%Y %H:%M").date(),
            lambda current: datetime.strptime(current, "%d-%m-%Y").date(),
            lambda current: datetime.strptime(current, "%d-%m-%Y %H:%M:%S").date(),
            lambda current: datetime.strptime(current, "%d-%m-%Y %H:%M").date(),
        ):
            try:
                return parser(cleaned)
            except ValueError:
                continue

        try:
            excel_serial = float(cleaned)
            excel_base = datetime(1899, 12, 30, tzinfo=UTC)
            return (excel_base + timedelta(days=excel_serial)).date()
        except (ValueError, OverflowError):
            return None

    def _date_requires_manual_review(self, *, raw_date: str, column_name: str | None) -> bool:
        if not self._is_ambiguous_slash_date(raw_date):
            return False

        normalized_column = self._normalize_column_name(column_name or "")
        if "fecha" in normalized_column:
            return False

        return normalized_column in {
            "date",
            "startdate",
            "transactiondate",
            "bookingdate",
            "posteddate",
            "valuedate",
        }

    def _is_ambiguous_slash_date(self, value: str) -> bool:
        match = re.fullmatch(
            r"\s*(\d{1,2})/(\d{1,2})/(\d{4})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\s*",
            value,
        )
        if match is None:
            return False

        first, second = (int(part) for part in match.groups()[:2])
        return first <= 12 and second <= 12 and first != second

    def _normalize_human_date(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        month_aliases = {
            "ene": "01",
            "enero": "01",
            "feb": "02",
            "febrero": "02",
            "mar": "03",
            "marzo": "03",
            "abr": "04",
            "abril": "04",
            "may": "05",
            "mayo": "05",
            "jun": "06",
            "junio": "06",
            "jul": "07",
            "julio": "07",
            "ago": "08",
            "agosto": "08",
            "sep": "09",
            "sept": "09",
            "septiembre": "09",
            "oct": "10",
            "octubre": "10",
            "nov": "11",
            "noviembre": "11",
            "dic": "12",
            "diciembre": "12",
        }
        match = re.fullmatch(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})", normalized)
        if match:
            day, month_label, year = match.groups()
            month_number = month_aliases.get(month_label)
            if month_number:
                return f"{int(day):02d} {month_number} {year}"
        return value.strip()

    def _parse_amount(self, value: str) -> Decimal | None:
        if not value:
            return None

        cleaned = value.strip().replace(" ", "")
        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")

        cleaned = cleaned.replace("€", "").replace("$", "").replace("£", "")

        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    def _normalize_column_name(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        return "".join(
            char for char in normalized if char.isalnum() and not unicodedata.combining(char)
        )

    def _decode_csv_content(self, content: bytes) -> str:
        decode_attempts = ("utf-8-sig", "utf-8", "latin-1")
        last_error: UnicodeDecodeError | None = None
        for encoding in decode_attempts:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError as exc:
                last_error = exc

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded CSV could not be decoded with a supported text encoding",
        ) from last_error

    def _normalize_description_key(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        normalized = re.sub(r"\b\d+(?:[.,/-]\d+)*\b", " ", normalized)
        normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
        tokens = [token for token in normalized.split() if token]
        return " ".join(tokens)

    def _merchant_pattern_key(self, value: str) -> str:
        tokens = [
            token
            for token in self._normalize_description_key(value).split()
            if len(token) > 2 and token not in DESCRIPTION_STOPWORDS
        ]
        if not tokens:
            return ""
        return tokens[0]

    def _to_cell_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()

    def _require_account(self, *, user_id: uuid.UUID, account_id: uuid.UUID) -> Account:
        account = self.account_repository.get_for_user(user_id=user_id, account_id=account_id)
        if account is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The selected account does not exist for the current user",
            )
        return account

    def _require_category_if_present(
        self,
        *,
        user_id: uuid.UUID,
        category_id: uuid.UUID | None,
    ) -> Category | None:
        if category_id is None:
            return None

        category = self.category_repository.get_for_user(user_id=user_id, category_id=category_id)
        if category is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The selected category does not exist for the current user",
            )
        return category

    def _existing_import_fingerprints(
        self,
        *,
        user_id: uuid.UUID,
        items: list[TransactionImportCommitItem],
    ) -> set[TransactionImportFingerprint]:
        if not items:
            return set()

        date_from = min(item.date for item in items)
        date_to = max(item.date for item in items)
        account_ids = {item.account_id for item in items}

        fingerprints: set[TransactionImportFingerprint] = set()
        for account_id in account_ids:
            transactions = self.repository.list_all_for_user(
                user_id=user_id,
                account_id=account_id,
                date_from=date_from,
                date_to=date_to,
            )
            for transaction in transactions:
                fingerprints.add(
                    self._transaction_fingerprint(
                        account_id=transaction.account_id,
                        date_value=transaction.date,
                        amount=transaction.amount,
                        currency=transaction.currency,
                        description=transaction.description,
                        notes=transaction.notes,
                    )
                )
        return fingerprints

    def _transaction_fingerprint(
        self,
        *,
        account_id: uuid.UUID,
        date_value: date,
        amount: Decimal,
        currency: str,
        description: str,
        notes: str | None,
    ) -> TransactionImportFingerprint:
        return TransactionImportFingerprint(
            account_id=account_id,
            date=date_value,
            amount=amount,
            currency=currency.strip().upper(),
            description=" ".join(description.split()).casefold(),
            notes=" ".join((notes or "").split()).casefold(),
        )
